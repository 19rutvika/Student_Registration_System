from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl , xlrd
from openpyxl import Workbook
import pathlib

background="#06283D"
framebg="#EDEDED"
framefg="#06283D"

root=Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)

###### icon
image_icon=PhotoImage(file="Images/icon.png")
root.iconphoto(False,image_icon)

file=pathlib.Path("C:/Users/Ritz/IT_PROJECTS/PYTHON/STUDENT REGISTRATION SYSTEM/Student_data.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1']="Name"
    sheet['C1']="Class"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date of Registartion"
    sheet['G1']="Religion"
    sheet['H1']="Skill"
    sheet['I1']="Father Name"
    sheet['J1']="Mother Name"
    sheet['K1']="Father's Occupation"
    sheet['L1']="Mother's Occupation"

    file.save('C:/Users/Ritz/IT_PROJECTS/PYTHON/STUDENT REGISTRATION SYSTEM/Student_data.xlsx')

    

#Exit window
def Exit():
    root.destroy()

######showimage
def showimage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
                                        title="Select image file",
                                        filetype=(("JPG File","*.jpg"),
                                                  ("PNG File","*.png"),
                                                  ("All Files","*.txt")))
    img=(Image.open(filename))
    resized_image = img.resize((190, 194), Image.LANCZOS)

    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

#Registration No.
#now each time we have to enter registartion no.
#lets design automatic registration no.

def registration_no():
    file=openpyxl.load_workbook("C:/Users/Ritz/IT_PROJECTS/PYTHON/STUDENT REGISTRATION SYSTEM/Student_data.xlsx")
    sheet=file.active
    row=sheet.max_row
    max_row_value=sheet.cell(row=row,column=1).value
    

    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set("1")

#clear
def Clear():
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skill.set('')
    F_Name.set('')
    M_Name.set('')
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Class.set("Select Class")

    registration_no()

    saveButton.config(state='normal')
    img1=PhotoImage(file='Images/upload photo.png')
    lbl.config(image=img1)
    lbl.image=img1

    img=""

def Save():
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    try:
        G1=Gender
    except:
        messagebox.showerror("error","Select Gender!")

    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=Skill.get()
    fathername=F_Name.get()
    mothername=M_Name.get()
    F1=Father_Occupation.get()
    M1=Mother_Occupation.get()

    if N1=="" or C1=="Select Class" or D2=="" or Re1=="" or S1=="" or fathername=="" or mothername=="" or F1=="" or M1=="":
        messagebox.showerror("error","Few Data is missing!")
    else:
        file=openpyxl.load_workbook("C:/Users/Ritz/IT_PROJECTS/PYTHON/STUDENT REGISTRATION SYSTEM/Student_data.xlsx")
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=C1)
        sheet.cell(column=4,row=sheet.max_row,value=G1)
        sheet.cell(column=5,row=sheet.max_row,value=D2)
        sheet.cell(column=6,row=sheet.max_row,value=D1)
        sheet.cell(column=7,row=sheet.max_row,value=Re1)
        sheet.cell(column=8,row=sheet.max_row,value=S1)
        sheet.cell(column=9,row=sheet.max_row,value=fathername)
        sheet.cell(column=10,row=sheet.max_row,value=mothername)
        sheet.cell(column=11,row=sheet.max_row,value=F1)
        sheet.cell(column=12,row=sheet.max_row,value=M1)

        file.save(r"C:/Users/Ritz/IT_PROJECTS/PYTHON/STUDENT REGISTRATION SYSTEM/Student_data.xlsx")

        try:
            img.save("Student Images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","Profile Picture is not availabel!!!")

        messagebox.showinfo("info","Sucessfully data entered!!!")
        Clear()
        registration_no()

            

#gender
def selection():
    global Gender
    value=radio.get()
    if value==1:
        Gender="Male"
    else:
        Gender="Female"

#top frames
Label(root,text="Email: rutvikaradadiya19@gmail.com",width=10,height=3,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRAION",width=10,height=2,bg="#c36464",fg='#fff',font='arial 25 bold').pack(side=TOP,fill=X)

#search box to update
Search=StringVar()
Entry(root,textvariable=Search,width=17,bd=2,font="arial 20").place(x=825,y=72)
imageicon3=PhotoImage(file="Images/search.png")
Srch=Button(root,text="Search",compound=LEFT,image=imageicon3,bg='#68ddfa',font="arial 13 bold")
Srch.place(x=1090,y=69)

imageicon4=PhotoImage(file="Images/Layer 4.png")
Update_Button=Button(root,image=imageicon4,bg='#c36464')
Update_Button.place(x=110,y=64)

#Registration and Date
Label(root,text="Registration No:",font="araial 13",fg=framebg,bg=background).place(x=30,y=150)
Label(root,text="Date:",font="araial 13",fg=framebg,bg=background).place(x=500,y=150)

Registration=IntVar()
Date=StringVar()

reg_entry = Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_entry.place(x=160,y=153)

registration_no() ####called here

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=153)
Date.set(d1)

#Student details
obj=LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date Of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

Label(obj,text="Class:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="Religion:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Skills:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

Name=StringVar()
name_entry= Entry(obj,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)

DOB=StringVar()
dob_entry= Entry(obj,textvariable=DOB,width=20,font="arial 10")
dob_entry.place(x=160,y=100)

radio=IntVar()
R1=Radiobutton(obj,text="Male", variable=radio, value=1, bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=150)
R1=Radiobutton(obj,text="Female", variable=radio, value=2, bg=framebg,fg=framefg,command=selection)
R1.place(x=200,y=150)

Religion=StringVar()
religion_entry= Entry(obj,textvariable=Religion,width=20,font="arial 10")
religion_entry.place(x=630,y=100)

Skill=StringVar()
skill_entry= Entry(obj,textvariable=Skill,width=20,font="arial 10")
skill_entry.place(x=630,y=150)

Class=Combobox(obj,values=['1','2','3','4','5','6','7','8','9','10','11','12'],font="Roboto 10",width=17,state="r")
Class.place(x=630,y=50)
Class.set("Select Class")




#Parents details
obj2=LabelFrame(root,text="Parent's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text="Father's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)

F_Name=StringVar()
f_entry= Entry(obj2,textvariable=F_Name,width=20,font="arial 10")
f_entry.place(x=160,y=50)

Father_Occupation=StringVar()
fo_entry= Entry(obj2,textvariable=Father_Occupation,width=20,font="arial 10")
fo_entry.place(x=160,y=100)

Label(obj2,text="Mother's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)

M_Name=StringVar()
m_entry= Entry(obj2,textvariable=M_Name,width=20,font="arial 10")
m_entry.place(x=630,y=50)

Mother_Occupation=StringVar()
mo_entry= Entry(obj2,textvariable=Mother_Occupation,width=20,font="arial 10")
mo_entry.place(x=630,y=100)


#image
f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file="Images/upload photo.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

#button

Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showimage).place(x=1000,y=370)

saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=Save)
saveButton.place(x=1000,y=450)

Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightpink",command=Clear).place(x=1000,y=530)

Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="grey",command=Exit).place(x=1000,y=610)


















root.mainloop()
